"""
ArtinSmartRealty V2 - Main API
FastAPI Application with Scheduling, Excel Export, and Background Tasks
"""

import os
import io
import asyncio
import secrets
import hashlib
import logging
from datetime import datetime, time, timedelta
from typing import Optional, List
from contextlib import asynccontextmanager

# Setup logger
logger = logging.getLogger(__name__)

from fastapi import FastAPI, HTTPException, Depends, Query, BackgroundTasks, Response, Header, UploadFile, File, Form, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, EmailStr, field_validator, field_serializer
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_, or_, delete
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
import openpyxl
from enum import Enum
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import jwt

from database import (
    init_db, async_session,
    Tenant, Lead, AgentAvailability, Appointment,
    TenantProperty, TenantProject, TenantKnowledge,
    LeadStatus, TransactionType, PropertyType, PaymentMethod, Purpose,
    AppointmentType, DayOfWeek, Language, ConversationState, SubscriptionStatus,
    get_leads_needing_reminder, get_appointments_needing_reminder,
    get_available_slots, book_slot, create_appointment, update_lead,
    get_db
)
from telegram_bot import bot_manager, handle_telegram_webhook
from whatsapp_bot import whatsapp_bot_manager, verify_webhook as verify_whatsapp_webhook
from roi_engine import generate_roi_pdf
from rate_limiter import rate_limit, rate_limiter, cleanup_rate_limiter
from security_headers import add_security_headers
from password_validator import validate_password_strength
from input_sanitizer import sanitize_text, sanitize_email, sanitize_phone

# Import API routers
from api import broadcast, catalogs, lotteries, admin, smart_upload
from auth_config import JWT_SECRET, JWT_ALGORITHM, JWT_EXPIRATION_HOURS, PASSWORD_SALT


# ==================== AUTH CONFIG ====================

security = HTTPBearer(auto_error=False)


def hash_password(password: str) -> str:
    """Hash password using PBKDF2 with SHA-256.
    
    Uses 600,000 iterations (OWASP 2023 recommendation).
    Previous: 100,000 iterations (too weak for modern GPUs).
    """
    return hashlib.pbkdf2_hmac('sha256', password.encode(), PASSWORD_SALT.encode(), 600000).hex()


def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password against hash using constant-time comparison.
    
    Prevents timing attacks where attacker can determine correct password
    characters by measuring response time differences.
    """
    computed_hash = hash_password(plain_password)
    # Use secrets.compare_digest for constant-time comparison
    return secrets.compare_digest(computed_hash, hashed_password)


def create_jwt_token(tenant_id: int, email: str) -> str:
    """Create JWT token for tenant."""
    from datetime import timezone
    payload = {
        "tenant_id": tenant_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def decode_jwt_token(token: str) -> dict:
    """Decode and verify JWT token."""
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def get_current_tenant(
    credentials: HTTPAuthorizationCredentials = Depends(security)
) -> Optional[Tenant]:
    """Get current tenant from JWT token."""
    if not credentials:
        return None
    
    payload = decode_jwt_token(credentials.credentials)
    tenant_id = payload.get("tenant_id")
    
    async with async_session() as session:
        result = await session.execute(select(Tenant).where(Tenant.id == tenant_id))
        return result.scalar_one_or_none()


async def verify_tenant_access(
    credentials: HTTPAuthorizationCredentials,
    tenant_id: int,
    db: AsyncSession
) -> Tenant:
    """
    Verify that the authenticated user has access to the given tenant.
    Super Admin (tenant_id=0) can access any tenant.
    Regular tenants can only access their own data.
    """
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    payload = decode_jwt_token(credentials.credentials)
    auth_tenant_id = payload.get("tenant_id")
    
    # Super Admin can access any tenant
    if auth_tenant_id == 0:
        result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
        tenant = result.scalar_one_or_none()
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant not found")
        return tenant
    
    # Regular tenant can only access their own data
    if auth_tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    return tenant


# ==================== AUTH MODELS ====================

class LoginRequest(BaseModel):
    email: EmailStr
    password: str = Field(..., min_length=6)


class RegisterRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    email: EmailStr
    password: str = Field(..., min_length=8, max_length=128, description="Password must be 8-128 characters with uppercase, lowercase, number, and special character")
    company_name: Optional[str] = Field(None, max_length=255)
    phone: Optional[str] = Field(None, max_length=50)


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    tenant_id: int
    name: str
    email: str
    is_super_admin: bool = False


class PasswordResetRequest(BaseModel):
    email: EmailStr


class PasswordResetConfirm(BaseModel):
    token: str
    new_password: str = Field(..., min_length=6)


# ==================== SUPER ADMIN CONFIG ====================
# Platform owner credentials (set these in .env for production)
SUPER_ADMIN_EMAIL = os.getenv("SUPER_ADMIN_EMAIL", "admin@artinsmartrealty.com")
SUPER_ADMIN_PASSWORD = os.getenv("SUPER_ADMIN_PASSWORD", "SuperAdmin123!")


# ==================== PYDANTIC MODELS ====================

class TenantCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    company_name: Optional[str] = Field(None, max_length=255)
    phone: Optional[str] = Field(None, max_length=50)
    email: Optional[str] = Field(None, max_length=255)
    telegram_bot_token: Optional[str] = Field(None, max_length=255)
    logo_url: Optional[str] = Field(None, max_length=512)
    primary_color: Optional[str] = Field("#D4AF37", max_length=20)
    # WhatsApp Business API fields
    whatsapp_phone_number_id: Optional[str] = Field(None, max_length=100)
    whatsapp_access_token: Optional[str] = Field(None, max_length=512)
    whatsapp_business_account_id: Optional[str] = Field(None, max_length=100)
    whatsapp_verify_token: Optional[str] = Field(None, max_length=255)


class TenantUpdate(BaseModel):
    """Model for partial tenant updates - all fields optional"""
    name: Optional[str] = Field(None, min_length=1, max_length=255)
    company_name: Optional[str] = Field(None, max_length=255)
    phone: Optional[str] = Field(None, max_length=50)
    email: Optional[str] = Field(None, max_length=255)
    telegram_bot_token: Optional[str] = Field(None, max_length=255)
    logo_url: Optional[str] = Field(None, max_length=512)
    primary_color: Optional[str] = Field(None, max_length=20)
    admin_chat_id: Optional[str] = Field(None, max_length=100)
    booking_url: Optional[str] = Field(None, max_length=512)
    contact_phone: Optional[str] = Field(None, max_length=50)
    whatsapp_link: Optional[str] = Field(None, max_length=512)
    # WhatsApp Business API fields
    whatsapp_phone_number_id: Optional[str] = Field(None, max_length=100)
    whatsapp_access_token: Optional[str] = Field(None, max_length=512)
    whatsapp_business_account_id: Optional[str] = Field(None, max_length=100)
    whatsapp_verify_token: Optional[str] = Field(None, max_length=255)


class TenantResponse(BaseModel):
    id: int
    name: str
    company_name: Optional[str]
    phone: Optional[str]
    email: Optional[str]
    telegram_bot_token: Optional[str]
    logo_url: Optional[str]
    primary_color: Optional[str]
    subscription_status: Optional[str]
    trial_ends_at: Optional[datetime]
    whatsapp_phone_number_id: Optional[str]
    whatsapp_access_token: Optional[str]
    whatsapp_business_account_id: Optional[str]
    whatsapp_verify_token: Optional[str]
    is_active: bool
    created_at: datetime

    class Config:
        from_attributes = True


class LeadResponse(BaseModel):
    id: int
    tenant_id: int
    name: Optional[str]
    phone: Optional[str]
    email: Optional[str]
    telegram_username: Optional[str]
    language: Optional[Language]
    status: Optional[LeadStatus]
    transaction_type: Optional[TransactionType]
    property_type: Optional[PropertyType]
    budget_min: Optional[float]
    budget_max: Optional[float]
    budget_currency: str
    payment_method: Optional[PaymentMethod]
    purpose: Optional[Purpose]
    bedrooms_min: Optional[int]
    bedrooms_max: Optional[int]
    preferred_location: Optional[str]
    taste_tags: List[str]
    voice_transcript: Optional[str]
    voice_entities: Optional[dict]
    image_description: Optional[str]
    image_search_results: Optional[int]
    conversation_state: Optional[ConversationState]
    source: str
    last_interaction: Optional[datetime]
    created_at: datetime
    
    # Sales Intelligence Fields
    lead_score: int = 0
    temperature: str = "cold"
    qr_scan_count: int = 0
    catalog_views: int = 0
    messages_count: int = 0
    total_interactions: int = 0

    class Config:
        from_attributes = True
    
    @field_serializer('language', 'status', 'transaction_type', 'property_type', 'payment_method', 'purpose', 'conversation_state')
    def serialize_all_enums_lowercase(self, value: Optional[Enum]) -> Optional[str]:
        """Convert ALL enum values to lowercase strings for API compatibility.
        
        Database stores enum NAMES (uppercase: 'FA', 'APARTMENT', 'BUY'),
        but API expects enum VALUES (lowercase: 'fa', 'apartment', 'buy').
        """
        if value is None:
            return None
        if isinstance(value, Enum):
            # Use .value if available (lowercase), else .name.lower()
            return value.value if hasattr(value, 'value') else value.name.lower()
        return str(value).lower() if value else None


class LeadUpdate(BaseModel):
    """Validated lead update model - only allow specific fields to be updated."""
    name: Optional[str] = Field(None, max_length=255)
    phone: Optional[str] = Field(None, max_length=50)
    status: Optional[LeadStatus] = None
    transaction_type: Optional[TransactionType] = None
    property_type: Optional[PropertyType] = None
    budget_min: Optional[float] = Field(None, ge=0)
    budget_max: Optional[float] = Field(None, ge=0)
    payment_method: Optional[PaymentMethod] = None
    purpose: Optional[Purpose] = None
    taste_tags: Optional[List[str]] = None
    notes: Optional[str] = None


class ScheduleSlotCreate(BaseModel):
    day_of_week: str  # Accept string, will convert to enum in endpoint
    start_time: str = Field(..., description="Time in HH:MM format")
    end_time: str = Field(..., description="Time in HH:MM format")
    
    @field_validator('day_of_week')
    @classmethod
    def validate_day_of_week(cls, v: str) -> str:
        """Validate and normalize day of week to lowercase."""
        if isinstance(v, str):
            v = v.lower()
        # Verify it's a valid day
        valid_days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        if v not in valid_days:
            raise ValueError(f'day_of_week must be one of: {", ".join(valid_days)}')
        return v


class ScheduleSlotsRequest(BaseModel):
    """Request body for creating multiple schedule slots."""
    slots: List[ScheduleSlotCreate]


class ScheduleSlotResponse(BaseModel):
    id: int
    day_of_week: DayOfWeek
    start_time: str
    end_time: str
    is_booked: bool

    class Config:
        from_attributes = True


class AppointmentCreate(BaseModel):
    lead_id: int
    appointment_type: AppointmentType
    scheduled_date: datetime
    duration_minutes: int = 60
    location: Optional[str] = None
    meeting_link: Optional[str] = None
    property_reference: Optional[str] = None
    notes: Optional[str] = None


class AppointmentResponse(BaseModel):
    id: int
    lead_id: int
    appointment_type: AppointmentType
    scheduled_date: datetime
    duration_minutes: int
    location: Optional[str]
    meeting_link: Optional[str]
    is_confirmed: bool
    is_cancelled: bool
    created_at: datetime

    class Config:
        from_attributes = True


class DashboardStats(BaseModel):
    total_leads: int
    active_deals: int
    qualified_leads: int
    scheduled_viewings: int
    conversion_rate: float
    leads_by_status: dict
    leads_by_purpose: dict


# ==================== TENANT DATA MODELS ====================

class TenantPropertyCreate(BaseModel):
    """Model for creating/updating tenant properties."""
    name: str = Field(..., min_length=1, max_length=255)
    property_type: PropertyType
    transaction_type: Optional[TransactionType] = None
    location: str = Field(..., min_length=1, max_length=255)
    address: Optional[str] = None
    price: Optional[float] = Field(None, ge=0)
    price_per_sqft: Optional[float] = Field(None, ge=0)
    currency: str = Field("AED", max_length=10)
    bedrooms: Optional[int] = Field(None, ge=0)
    bathrooms: Optional[int] = Field(None, ge=0)
    area_sqft: Optional[float] = Field(None, ge=0)
    features: Optional[List[str]] = []
    description: Optional[str] = None
    full_description: Optional[str] = None  # Rich formatted description with emojis
    expected_roi: Optional[float] = Field(None, ge=0, le=100)
    rental_yield: Optional[float] = Field(None, ge=0, le=100)
    golden_visa_eligible: bool = False
    is_urgent: bool = False  # For "Urgent Sale" properties
    images: Optional[List[str]] = []
    image_urls: Optional[List[str]] = []  # Added for frontend compatibility
    is_featured: bool = False
    
    model_config = {"extra": "ignore"}  # Ignore unknown fields from frontend
    
    @field_validator('image_urls', 'images', mode='before')
    @classmethod
    def convert_image_objects_to_urls(cls, v):
        """Convert image objects from frontend to URL strings."""
        if not v:
            return []
        
        urls = []
        for item in v:
            if isinstance(item, dict):
                # Frontend sends objects with 'url' field
                urls.append(item.get('url', ''))
            elif isinstance(item, str):
                urls.append(item)
        
        return urls


class TenantPropertyResponse(BaseModel):
    id: int
    name: str
    property_type: PropertyType
    transaction_type: Optional[TransactionType] = None
    location: str
    price: Optional[float]
    bedrooms: Optional[int]
    features: Optional[List[str]]
    description: Optional[str]
    full_description: Optional[str]
    golden_visa_eligible: bool = False
    is_available: bool = True
    is_featured: bool = False
    is_urgent: bool = False
    image_urls: Optional[List[str]]
    primary_image: Optional[str]
    image_files: Optional[List[dict]]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True
    
    @field_serializer('property_type', 'transaction_type')
    def serialize_enums_lowercase(self, value: Optional[Enum]) -> Optional[str]:
        """Convert enum values to lowercase strings for API/database compatibility.
        
        Database stores enum NAMES (uppercase: 'APARTMENT', 'BUY'),
        but API expects enum VALUES (lowercase: 'apartment', 'buy').
        """
        if value is None:
            return None
        if isinstance(value, Enum):
            return value.value if hasattr(value, 'value') else value.name.lower()
        return str(value).lower() if value else None


class TenantProjectCreate(BaseModel):
    """Model for creating/updating tenant off-plan projects."""
    name: str = Field(..., min_length=1, max_length=255)
    developer: Optional[str] = Field(None, max_length=255)
    location: str = Field(..., min_length=1, max_length=255)
    starting_price: Optional[float] = Field(None, ge=0)
    price_per_sqft: Optional[float] = Field(None, ge=0)
    currency: str = Field("AED", max_length=10)
    payment_plan: Optional[str] = Field(None, max_length=255)
    down_payment_percent: Optional[float] = Field(None, ge=0, le=100)
    handover_date: Optional[datetime] = None
    completion_percent: Optional[float] = Field(None, ge=0, le=100)
    projected_roi: Optional[float] = Field(None, ge=0, le=100)
    projected_rental_yield: Optional[float] = Field(None, ge=0, le=100)
    golden_visa_eligible: bool = False
    amenities: Optional[List[str]] = []
    unit_types: Optional[List[str]] = []
    description: Optional[str] = None
    selling_points: Optional[List[str]] = []
    is_featured: bool = False


class TenantProjectResponse(BaseModel):
    id: int
    name: str
    developer: Optional[str]
    location: str
    starting_price: Optional[float]
    payment_plan: Optional[str]
    golden_visa_eligible: bool
    is_active: bool
    is_featured: bool

    class Config:
        from_attributes = True


class TenantKnowledgeCreate(BaseModel):
    """Model for creating/updating tenant knowledge base entries."""
    category: str = Field(..., min_length=1, max_length=100)  # "faq", "policy", "service"
    title: str = Field(..., min_length=1, max_length=255)
    content: str = Field(..., min_length=1)
    language: Language = Language.EN
    keywords: Optional[List[str]] = []
    priority: int = Field(0, ge=0, le=100)


class TenantKnowledgeResponse(BaseModel):
    id: int
    category: str
    title: str
    content: str
    language: Language
    priority: int
    is_active: bool

    class Config:
        from_attributes = True


# ==================== SCHEDULER ====================

scheduler = AsyncIOScheduler()


async def ghost_protocol_job():
    """Background job for Ghost Protocol - send reminders to inactive leads."""
    try:
        await bot_manager.send_ghost_reminders()
    except Exception as e:
        print(f"Ghost protocol job error: {e}")


async def appointment_reminder_job():
    """Background job for appointment reminders."""
    try:
        await bot_manager.send_appointment_reminders()
    except Exception as e:
        print(f"Appointment reminder job error: {e}")


# ==================== LIFESPAN ====================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan - startup and shutdown."""
    # Startup
    print("🚀 Starting ArtinSmartRealty V2 - Unified Platform...")
    
    # Initialize database
    await init_db()
    print("✅ Database initialized")
    
    # Start Telegram bots for all active tenants
    async with async_session() as session:
        result = await session.execute(
            select(Tenant).where(
                Tenant.is_active == True,
                Tenant.telegram_bot_token.isnot(None)
            )
        )
        tenants = result.scalars().all()
        
        for tenant in tenants:
            try:
                await bot_manager.start_bot_for_tenant(tenant)
            except Exception as e:
                print(f"Failed to start bot for tenant {tenant.id}: {e}")
    
    # Start scheduler
    scheduler.add_job(
        ghost_protocol_job,
        trigger=IntervalTrigger(hours=1),
        id="ghost_protocol",
        replace_existing=True
    )
    scheduler.add_job(
        appointment_reminder_job,
        trigger=IntervalTrigger(hours=1),
        id="appointment_reminders",
        replace_existing=True
    )
    scheduler.start()
    print("✅ Background scheduler started")
    
    # Start Morning Coffee Report scheduler in bot_manager
    await bot_manager.start_scheduler()
    print("✅ Morning Coffee Report scheduler started")
    
    # 🆕 Start Unified Follow-up Engine
    from followup_engine import start_followup_engine
    await start_followup_engine()
    print("✅ Unified Follow-up Engine started")
    
    # Start rate limiter cleanup task
    asyncio.create_task(cleanup_rate_limiter())
    print("✅ Rate limiter cleanup task started")
    
    yield
    
    # Shutdown
    print("🛑 Shutting down...")
    scheduler.shutdown()
    await bot_manager.stop_scheduler()
    await bot_manager.stop_all_bots()
    
    # Stop Follow-up Engine
    from followup_engine import stop_followup_engine
    await stop_followup_engine()
    
    print("✅ Shutdown complete")


# ==================== FASTAPI APP ====================

app = FastAPI(
    title="ArtinSmartRealty V2 API",
    description="Multi-Tenant Real Estate SaaS Platform",
    version="2.0.0",
    lifespan=lifespan
)

# Add exception handler for Pydantic validation errors
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
import json

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    # Log validation errors with request details
    logger.error(f"❌ Validation Error on {request.method} {request.url}")
    logger.error(f"❌ Errors: {exc.errors()}")
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors()}
    )

# CORS Middleware - Use environment variable for allowed origins in production
# Security: NEVER use wildcard "*" with credentials enabled (CSRF risk)
ALLOWED_ORIGINS_DEFAULT = [
    "http://localhost:3000",
    "http://localhost:5173",
    "http://127.0.0.1:3000",
]
CORS_ORIGINS = os.getenv("CORS_ORIGINS", ",".join(ALLOWED_ORIGINS_DEFAULT)).split(",")

# Security check: Prevent wildcard with credentials
if "*" in CORS_ORIGINS:
    logger.error(
        "SECURITY ERROR: CORS wildcard '*' detected with credentials enabled! "
        "This allows ANY website to steal user data via CSRF. "
        "Set CORS_ORIGINS in .env to specific domains only."
    )
    # In production, raise error. In dev, allow but warn
    if os.getenv("ENVIRONMENT", "development") == "production":
        raise RuntimeError("CORS wildcard not allowed in production!")
    else:
        logger.warning("⚠️  DEVELOPMENT MODE: Allowing CORS wildcard (NOT FOR PRODUCTION)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add security headers middleware
add_security_headers(app)

# Mount static files for uploads
UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/app/uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Include API routers
app.include_router(admin.router, prefix="/api")
app.include_router(broadcast.router)
app.include_router(catalogs.router)
app.include_router(lotteries.router)
app.include_router(smart_upload.router)  # 🚀 Smart PDF/Image Upload with AI

# 🆕 Include Unified Lead Management API
from api.unified_routes import router as unified_router
app.include_router(unified_router)

# 🆕 Include LinkedIn Scraper Integration API (PRO PLAN)
from api.linkedin_routes import router as linkedin_router
app.include_router(linkedin_router)

# 🤖 Include Follow-up Management API
from api.followup_routes import router as followup_router
app.include_router(followup_router)

# 💳 Include Subscription & Registration API
from api.subscription import router as subscription_router
app.include_router(subscription_router)

# 🔒 Include Super Admin Subscription Management API
from api.admin_subscription import router as admin_subscription_router
app.include_router(admin_subscription_router)

# 🏥 Include Health Check API
from api.health import router as health_router
app.include_router(health_router)


# ==================== LEGACY HEALTH CHECK (DEPRECATED) ====================

@app.get("/health")
async def health_check():
    """Basic health check endpoint."""
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@app.get("/admin/health/dashboard")
async def health_dashboard(current_tenant: Tenant = Depends(get_current_tenant)):
    """
    Comprehensive health dashboard for platform owner.
    Shows status of all critical services and metrics.
    Only accessible by super admin.
    """
    # Security: Only super admin can access
    if not current_tenant or not current_tenant.is_super_admin:
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    health_status = {
        "timestamp": datetime.utcnow().isoformat(),
        "services": {},
        "metrics": {},
        "errors": []
    }
    
    # Check Database
    try:
        async with async_session() as session:
            result = await session.execute(select(Tenant).limit(1))
            result.scalar_one_or_none()
        health_status["services"]["database"] = {
            "status": "healthy",
            "type": "PostgreSQL"
        }
    except Exception as e:
        health_status["services"]["database"] = {
            "status": "unhealthy",
            "error": str(e)
        }
        health_status["errors"].append(f"Database: {str(e)}")
    
    # Check Redis
    try:
        from redis_manager import redis_manager
        if redis_manager.redis_client:
            await redis_manager.redis_client.ping()
            health_status["services"]["redis"] = {"status": "healthy"}
        else:
            health_status["services"]["redis"] = {
                "status": "degraded",
                "message": "Not connected - bot running in degraded mode"
            }
    except Exception as e:
        health_status["services"]["redis"] = {
            "status": "unhealthy",
            "error": str(e)
        }
        health_status["errors"].append(f"Redis: {str(e)}")
    
    # Check Gemini API
    try:
        import google.generativeai as genai
        if GEMINI_API_KEY := os.getenv("GEMINI_API_KEY"):
            genai.configure(api_key=GEMINI_API_KEY)
            model = genai.GenerativeModel('gemini-2.0-flash-exp')
            # Quick test
            test_response = model.generate_content("Hello")
            if test_response.text:
                health_status["services"]["gemini_api"] = {"status": "healthy"}
            else:
                health_status["services"]["gemini_api"] = {
                    "status": "degraded",
                    "message": "API responding but empty response"
                }
        else:
            health_status["services"]["gemini_api"] = {
                "status": "unconfigured",
                "message": "GEMINI_API_KEY not set"
            }
            health_status["errors"].append("Gemini API: Missing API key")
    except Exception as e:
        health_status["services"]["gemini_api"] = {
            "status": "unhealthy",
            "error": str(e)
        }
        health_status["errors"].append(f"Gemini API: {str(e)}")
    
    # Get metrics
    try:
        async with async_session() as session:
            # Total tenants
            tenant_count = await session.execute(select(Tenant))
            health_status["metrics"]["total_tenants"] = len(tenant_count.scalars().all())
            
            # Active bots (tenants with bot tokens)
            active_bots = await session.execute(
                select(Tenant).where(Tenant.telegram_bot_token.isnot(None))
            )
            health_status["metrics"]["active_telegram_bots"] = len(active_bots.scalars().all())
            
            # Total leads
            total_leads = await session.execute(select(Lead))
            health_status["metrics"]["total_leads"] = len(total_leads.scalars().all())
            
            # Leads today
            today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
            leads_today = await session.execute(
                select(Lead).where(Lead.created_at >= today_start)
            )
            health_status["metrics"]["leads_today"] = len(leads_today.scalars().all())
    except Exception as e:
        health_status["errors"].append(f"Metrics collection: {str(e)}")
    
    # Overall status
    if health_status["errors"]:
        health_status["overall_status"] = "degraded" if health_status["services"].get("database", {}).get("status") == "healthy" else "unhealthy"
    else:
        health_status["overall_status"] = "healthy"
    
    return health_status


# ==================== AUTHENTICATION ENDPOINTS ====================

@app.post("/api/auth/register", response_model=LoginResponse)
async def register(data: RegisterRequest, db: AsyncSession = Depends(get_db)):
    """
    Register a new tenant/agent.
    DISABLED for public registration - use Super Admin panel instead.
    To enable public registration, remove the HTTPException below.
    """
    # Disable public registration - only admin can create tenants
    raise HTTPException(
        status_code=403,
        detail="Public registration is disabled. Please contact admin for account creation."
    )
    
    # Uncomment below to enable public registration:
    """
    # Validate password strength
    validate_password_strength(data.password, min_length=8)
    
    # Sanitize inputs
    email = sanitize_email(data.email)
    name = sanitize_text(data.name, max_length=255)
    
    # Check if email already exists
    result = await db.execute(select(Tenant).where(Tenant.email == email))
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Sanitize optional fields
    company_name = sanitize_text(data.company_name, max_length=255) if data.company_name else None
    phone = sanitize_phone(data.phone) if data.phone else None
    
    # Create new tenant
    tenant = Tenant(
        name=name,
        email=email,
        password_hash=hash_password(data.password),
        company_name=company_name,
        phone=phone,
        trial_ends_at=datetime.utcnow() + timedelta(days=14)  # 14-day trial
    )
    db.add(tenant)
    await db.commit()
    await db.refresh(tenant)
    
    # Generate JWT token
    token = create_jwt_token(tenant.id, tenant.email)
    
    return LoginResponse(
        access_token=token,
        tenant_id=tenant.id,
        name=tenant.name,
        email=tenant.email
    )
    """


@app.post("/api/auth/login", response_model=LoginResponse)
async def login(request: Request, data: LoginRequest, db: AsyncSession = Depends(get_db)):
    """Login with email and password. Supports both tenants and super admin.
    Rate limited: 5 attempts per minute to prevent brute force attacks.
    """
    
    # Apply rate limit
    client_ip = request.client.host if request.client else "unknown"
    if "X-Forwarded-For" in request.headers:
        client_ip = request.headers["X-Forwarded-For"].split(",")[0].strip()
    
    is_limited, retry_after = rate_limiter.is_rate_limited(
        client_ip, "/api/auth/login", max_requests=5, window_seconds=60
    )
    
    if is_limited:
        raise HTTPException(
            status_code=429,
            detail=f"Too many login attempts. Try again in {retry_after} seconds.",
            headers={"Retry-After": str(retry_after)}
        )
    
    # Check for Super Admin login first (constant-time comparison)
    email_match = secrets.compare_digest(data.email, SUPER_ADMIN_EMAIL)
    password_match = secrets.compare_digest(data.password, SUPER_ADMIN_PASSWORD)
    
    if email_match and password_match:
        # Super admin uses tenant_id 0 to indicate admin status
        token = create_jwt_token(0, SUPER_ADMIN_EMAIL)
        return LoginResponse(
            access_token=token,
            tenant_id=0,
            name="Super Admin",
            email=SUPER_ADMIN_EMAIL,
            is_super_admin=True
        )
    
    # Normal tenant login
    result = await db.execute(select(Tenant).where(Tenant.email == data.email))
    tenant = result.scalar_one_or_none()
    
    if not tenant or not tenant.password_hash:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not verify_password(data.password, tenant.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not tenant.is_active:
        raise HTTPException(status_code=403, detail="Account is disabled")
    
    # Generate JWT token
    token = create_jwt_token(tenant.id, tenant.email)
    
    return LoginResponse(
        access_token=token,
        tenant_id=tenant.id,
        name=tenant.name,
        email=tenant.email,
        is_super_admin=False
    )


@app.post("/api/auth/forgot-password")
async def forgot_password(request: Request, data: PasswordResetRequest, db: AsyncSession = Depends(get_db)):
    """Request password reset token.
    Rate limited: 3 attempts per hour to prevent email bombing.
    """
    
    # Apply rate limit
    client_ip = request.client.host if request.client else "unknown"
    if "X-Forwarded-For" in request.headers:
        client_ip = request.headers["X-Forwarded-For"].split(",")[0].strip()
    
    is_limited, retry_after = rate_limiter.is_rate_limited(
        client_ip, "/api/auth/forgot-password", max_requests=3, window_seconds=3600
    )
    
    if is_limited:
        raise HTTPException(
            status_code=429,
            detail=f"Too many password reset attempts. Try again in {retry_after // 60} minutes.",
            headers={"Retry-After": str(retry_after)}
        )
    
    result = await db.execute(select(Tenant).where(Tenant.email == data.email))
    tenant = result.scalar_one_or_none()
    
    # Always return success to prevent email enumeration
    if not tenant:
        return {"message": "If the email exists, a reset link has been sent"}
    
    # Generate reset token
    reset_token = secrets.token_urlsafe(32)
    tenant.reset_token = reset_token
    tenant.reset_token_expires = datetime.utcnow() + timedelta(hours=1)
    await db.commit()
    
    # In production, send email with reset link
    # TODO: Integrate email service (SendGrid, SES, etc.)
    # Example: send_email(tenant.email, "Password Reset", f"Reset link: {FRONTEND_URL}/reset?token={reset_token}")
    
    # For development only - in production, NEVER return the token
    if os.getenv("DEBUG", "false").lower() == "true":
        return {"message": "If the email exists, a reset link has been sent", "_debug_token": reset_token}
    
    return {"message": "If the email exists, a reset link has been sent"}


@app.post("/api/auth/reset-password")
async def reset_password(data: PasswordResetConfirm, db: AsyncSession = Depends(get_db)):
    """Reset password using token."""
    result = await db.execute(
        select(Tenant).where(
            Tenant.reset_token == data.token,
            Tenant.reset_token_expires > datetime.utcnow()
        )
    )
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
    
    # Update password
    tenant.password_hash = hash_password(data.new_password)
    tenant.reset_token = None
    tenant.reset_token_expires = None
    await db.commit()
    
    return {"message": "Password reset successful"}


# ==================== SUPER ADMIN ENDPOINTS ====================

async def verify_super_admin(credentials: HTTPAuthorizationCredentials):
    """Verify that the request is from super admin."""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        tenant_id = payload.get("tenant_id")
        
        logger.info(f"🔐 Super admin verification - tenant_id: {tenant_id}, email: {payload.get('email')}")
        
        if tenant_id != 0:  # Super admin has tenant_id = 0
            logger.warning(f"❌ Non-admin access attempt - tenant_id: {tenant_id}")
            raise HTTPException(status_code=403, detail="Super admin access required")
        
        logger.info("✅ Super admin verified successfully")
        return payload
    except HTTPException:
        raise  # Re-raise HTTPException as-is
    except Exception as e:
        logger.error(f"❌ Super admin verification failed: {type(e).__name__}: {e}")
        raise HTTPException(status_code=403, detail="Super admin access required")


@app.get("/api/admin/tenants")
async def list_all_tenants(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Get all tenants. Super admin only."""
    await verify_super_admin(credentials)
    
    result = await db.execute(select(Tenant).order_by(Tenant.created_at.desc()))
    tenants = result.scalars().all()
    
    return [
        {
            "id": t.id,
            "name": t.name,
            "email": t.email,
            "company_name": t.company_name,
            "phone": t.phone,
            "telegram_bot_token": t.telegram_bot_token,
            "subscription_status": t.subscription_status.value if t.subscription_status else "TRIAL",
            "is_active": t.is_active,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "trial_ends_at": t.trial_ends_at.isoformat() if t.trial_ends_at else None
        }
        for t in tenants
    ]


@app.get("/api/admin/tenants/{tenant_id}/credentials")
async def get_tenant_credentials(
    tenant_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Get tenant login credentials (email only - password cannot be retrieved). Super admin only."""
    await verify_super_admin(credentials)
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    return {
        "id": tenant.id,
        "name": tenant.name,
        "email": tenant.email,
        "note": "Password is hashed and cannot be retrieved. Use /reset-password endpoint to set a new password."
    }


@app.post("/api/admin/tenants")
async def create_tenant(
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Create a new tenant. Super admin only."""
    await verify_super_admin(credentials)
    
    # Check if email already exists
    result = await db.execute(select(Tenant).where(Tenant.email == data.get("email")))
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create tenant
    tenant = Tenant(
        name=data["name"],
        email=data["email"],
        password_hash=hash_password(data["password"]),
        company_name=data.get("company_name"),
        phone=data.get("phone"),
        telegram_bot_token=data.get("telegram_bot_token"),
        subscription_status=SubscriptionStatus[data.get("subscription_status", "TRIAL")],
        is_active=True,
        created_at=datetime.utcnow(),
        trial_ends_at=datetime.utcnow() + timedelta(days=14)  # 14 day trial
    )
    
    db.add(tenant)
    await db.commit()
    await db.refresh(tenant)
    
    return {"id": tenant.id, "message": "Tenant created successfully"}


@app.put("/api/admin/tenants/{tenant_id}")
async def update_tenant(
    tenant_id: int,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Update a tenant. Super admin only."""
    await verify_super_admin(credentials)
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    # Update fields
    if "name" in data:
        tenant.name = data["name"]
    if "email" in data:
        tenant.email = data["email"]
    if "company_name" in data:
        tenant.company_name = data["company_name"]
    if "phone" in data:
        tenant.phone = data["phone"]
    if "telegram_bot_token" in data:
        tenant.telegram_bot_token = data["telegram_bot_token"]
    if "subscription_status" in data:
        tenant.subscription_status = SubscriptionStatus[data["subscription_status"]]
    if "is_active" in data:
        tenant.is_active = data["is_active"]
    
    tenant.updated_at = datetime.utcnow()
    await db.commit()
    
    return {"message": "Tenant updated successfully"}


@app.delete("/api/admin/tenants/{tenant_id}")
async def delete_tenant(
    tenant_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Delete a tenant. Super admin only."""
    await verify_super_admin(credentials)
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    await db.delete(tenant)
    await db.commit()
    
    return {"message": "Tenant deleted successfully"}


@app.post("/api/admin/tenants/{tenant_id}/reset-password")
async def reset_tenant_password(
    tenant_id: int,
    new_password: str = Body(..., embed=True, min_length=8),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Reset tenant password. Super admin only."""
    await verify_super_admin(credentials)
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    # Hash new password
    tenant.password_hash = hash_password(new_password)
    tenant.updated_at = datetime.utcnow()
    await db.commit()
    
    logger.info(f"🔐 Password reset for tenant {tenant.name} (ID: {tenant_id})")
    
    return {
        "message": "Password reset successfully",
        "tenant_email": tenant.email,
        "new_password": new_password  # Return it so admin can share with tenant
    }


@app.get("/api/auth/me")
async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Get current authenticated tenant or super admin info."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    payload = decode_jwt_token(credentials.credentials)
    tenant_id = payload.get("tenant_id")
    
    # Super Admin case
    if tenant_id == 0:
        return {
            "id": 0,
            "name": "Super Admin",
            "email": SUPER_ADMIN_EMAIL,
            "is_super_admin": True,
            "is_active": True,
            "subscription_status": "active",
            "created_at": datetime.utcnow()
        }
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    return tenant


# ==================== TENANT ENDPOINTS ====================

@app.post("/api/tenants", response_model=TenantResponse)
async def create_tenant(tenant_data: TenantCreate, db: AsyncSession = Depends(get_db)):
    """Create a new tenant (agent)."""
    tenant = Tenant(**tenant_data.model_dump())
    db.add(tenant)
    await db.commit()
    await db.refresh(tenant)
    
    # Start bot if token provided
    if tenant.telegram_bot_token:
        try:
            await bot_manager.start_bot_for_tenant(tenant)
        except Exception as e:
            print(f"Failed to start bot: {e}")
    
    return tenant


@app.get("/api/tenants/{tenant_id}", response_model=TenantResponse)
async def get_tenant(
    tenant_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Get tenant by ID. Requires authentication."""
    await verify_tenant_access(credentials, tenant_id, db)
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    return tenant


@app.put("/api/tenants/{tenant_id}", response_model=TenantResponse)
async def update_tenant(
    tenant_id: int, 
    tenant_data: TenantUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Update tenant settings including bot tokens. Requires authentication."""
    await verify_tenant_access(credentials, tenant_id, db)
    
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    # Store old bot token before update to detect changes
    old_bot_token = tenant.telegram_bot_token
    
    # Update fields
    update_data = tenant_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if value is not None and hasattr(tenant, key):
            setattr(tenant, key, value)
    
    tenant.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(tenant)
    
    # Restart bot if token changed
    new_bot_token = tenant.telegram_bot_token
    if new_bot_token and new_bot_token != old_bot_token:
        try:
            await bot_manager.stop_bot_for_tenant(tenant_id)
            await bot_manager.start_bot_for_tenant(tenant)
        except Exception as e:
            print(f"Failed to restart bot: {e}")
    
    return tenant


@app.get("/api/tenants", response_model=List[TenantResponse])
async def list_tenants(
    skip: int = 0,
    limit: int = 100,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """List all tenants. Requires Super Admin authentication."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    payload = decode_jwt_token(credentials.credentials)
    auth_tenant_id = payload.get("tenant_id")
    
    # Only Super Admin (tenant_id=0) can list all tenants
    if auth_tenant_id != 0:
        raise HTTPException(status_code=403, detail="Super Admin access required")
    
    result = await db.execute(
        select(Tenant).offset(skip).limit(limit)
    )
    return result.scalars().all()


# ==================== LEAD ENDPOINTS ====================

@app.get("/api/tenants/{tenant_id}/leads/export")
async def export_leads_excel(
    tenant_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Export all leads to Excel (.xlsx) file. MUST be before /leads/{lead_id} route."""
    # Verify access
    await verify_tenant_access(credentials, tenant_id, db)
    
    # Get all leads for tenant
    result = await db.execute(
        select(Lead).where(Lead.tenant_id == tenant_id).order_by(Lead.created_at.desc())
    )
    leads = result.scalars().all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Define headers
    headers = [
        "ID", "Name", "Phone", "Telegram Username", "Language", "Status",
        "Transaction Type", "Property Type", "Budget Min", "Budget Max",
        "Payment Method", "Purpose", "Bedrooms Min", "Bedrooms Max",
        "Location", "Notes", "Source", "Created At"
    ]
    ws.append(headers)
    
    # Add data rows
    for lead in leads:
        ws.append([
            lead.id,
            lead.name or "",
            lead.phone or "",
            lead.telegram_username or "",
            lead.language.value if lead.language else "",
            lead.status.value if lead.status else "",
            lead.transaction_type.value if lead.transaction_type else "",
            lead.property_type.value if lead.property_type else "",
            lead.budget_min or "",
            lead.budget_max or "",
            lead.payment_method.value if lead.payment_method else "",
            lead.purpose.value if lead.purpose else "",
            lead.bedrooms_min or "",
            lead.bedrooms_max or "",
            lead.preferred_location or "",
            lead.notes or "",
            lead.source or "",
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else ""
        ])
    
    # Save to BytesIO
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=leads_{tenant_id}.xlsx"}
    )


@app.get("/api/tenants/{tenant_id}/leads", response_model=List[LeadResponse])
async def list_leads(
    tenant_id: int,
    status: Optional[LeadStatus] = None,
    purpose: Optional[Purpose] = None,
    skip: int = 0,
    limit: int = 100,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """List leads for a tenant with optional filtering. Requires authentication."""
    # Verify access
    await verify_tenant_access(credentials, tenant_id, db)
    
    query = select(Lead).where(Lead.tenant_id == tenant_id)
    
    if status:
        query = query.where(Lead.status == status)
    if purpose:
        query = query.where(Lead.purpose == purpose)
    
    # Order by lead_score DESC (NULL values last), then by created_at DESC
    # Use nullslast() to ensure NULLs appear at the end
    from sqlalchemy import nullslast
    query = query.order_by(
        nullslast(Lead.lead_score.desc()), 
        Lead.created_at.desc()
    ).offset(skip).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()


@app.get("/api/tenants/{tenant_id}/leads/{lead_id}", response_model=LeadResponse)
async def get_lead(tenant_id: int, lead_id: int, db: AsyncSession = Depends(get_db)):
    """Get a specific lead."""
    result = await db.execute(
        select(Lead).where(
            Lead.tenant_id == tenant_id,
            Lead.id == lead_id
        )
    )
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return lead


@app.patch("/api/tenants/{tenant_id}/leads/{lead_id}")
async def update_lead_endpoint(
    tenant_id: int,
    lead_id: int,
    updates: LeadUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update a lead with validated fields only."""
    result = await db.execute(
        select(Lead).where(
            Lead.tenant_id == tenant_id,
            Lead.id == lead_id
        )
    )
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Apply only non-None updates from validated model
    update_data = updates.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if value is not None and hasattr(lead, key):
            setattr(lead, key, value)
    
    lead.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(lead)
    
    return {"status": "updated", "lead_id": lead.id}


# ==================== SCHEDULING ENDPOINTS ====================

@app.get("/api/tenants/{tenant_id}/schedule", response_model=List[ScheduleSlotResponse])
async def list_schedule_slots(
    tenant_id: int,
    day_of_week: Optional[DayOfWeek] = None,
    available_only: bool = False,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """
    List schedule slots for a tenant. 
    Checks current week's bookings against Appointment table.
    """
    await verify_tenant_access(credentials, tenant_id, db)
    
    query = select(AgentAvailability).where(AgentAvailability.tenant_id == tenant_id)
    
    if day_of_week:
        query = query.where(AgentAvailability.day_of_week == day_of_week)
    
    result = await db.execute(query)
    slots = result.scalars().all()
    
    # Calculate is_booked for current week
    now = datetime.now()
    current_weekday = now.weekday()
    
    day_mapping = {
        'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
        'friday': 4, 'saturday': 5, 'sunday': 6
    }
    
    # 🚀 PERFORMANCE FIX: Fetch all appointments for this tenant ONCE (prevent N+1 queries)
    # Calculate date range for next 7 days to cover all possible slot occurrences
    date_start = now.date()
    date_end = date_start + timedelta(days=7)
    
    appointments_result = await db.execute(
        select(Appointment).where(
            Appointment.lead_id.in_(
                select(Lead.id).where(Lead.tenant_id == tenant_id)
            ),
            Appointment.scheduled_date >= datetime.combine(date_start, datetime.min.time()),
            Appointment.scheduled_date < datetime.combine(date_end, datetime.min.time()),
            Appointment.is_cancelled == False
        )
    )
    appointments = appointments_result.scalars().all()
    
    # Create a set of booked datetimes for O(1) lookup
    booked_datetimes = {appt.scheduled_date for appt in appointments}
    
    slot_responses = []
    for s in slots:
        target_weekday = day_mapping.get(s.day_of_week.value.lower(), 0)
        days_ahead = (target_weekday - current_weekday) % 7
        
        if days_ahead == 0:
            days_ahead = 7
        
        next_occurrence = now + timedelta(days=days_ahead)
        scheduled_datetime = datetime.combine(next_occurrence.date(), s.start_time)
        
        # 🚀 Check if booked using in-memory set (O(1) instead of database query)
        is_booked = scheduled_datetime in booked_datetimes
        
        # Filter if available_only requested
        if available_only and is_booked:
            continue
        
        slot_responses.append(ScheduleSlotResponse(
            id=s.id,
            day_of_week=s.day_of_week,
            start_time=s.start_time.strftime("%H:%M"),
            end_time=s.end_time.strftime("%H:%M"),
            is_booked=is_booked
        ))
    
    return slot_responses


@app.delete("/api/tenants/{tenant_id}/schedule/{slot_id}")
async def delete_schedule_slot(
    tenant_id: int,
    slot_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """
    Delete a schedule slot template.
    Note: This only deletes the recurring availability template.
    Existing appointments in Appointment table remain intact.
    """
    await verify_tenant_access(credentials, tenant_id, db)
    
    result = await db.execute(
        select(AgentAvailability).where(
            AgentAvailability.tenant_id == tenant_id,
            AgentAvailability.id == slot_id
        )
    )
    slot = result.scalar_one_or_none()
    
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    
    # Check if there are future appointments for this slot
    # (optional: warn if deleting a slot with future appointments)
    now = datetime.now()
    current_weekday = now.weekday()
    
    day_mapping = {
        'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
        'friday': 4, 'saturday': 5, 'sunday': 6
    }
    
    target_weekday = day_mapping.get(slot.day_of_week.value.lower(), 0)
    days_ahead = (target_weekday - current_weekday) % 7
    if days_ahead == 0:
        days_ahead = 7
    
    next_occurrence = now + timedelta(days=days_ahead)
    scheduled_datetime = datetime.combine(next_occurrence.date(), slot.start_time)
    
    future_appointments = await db.execute(
        select(Appointment).where(
            Appointment.lead_id.in_(
                select(Lead.id).where(Lead.tenant_id == tenant_id)
            ),
            Appointment.scheduled_date >= scheduled_datetime,
            Appointment.is_cancelled == False
        )
    )
    
    if future_appointments.scalar_one_or_none():
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete slot with future appointments. Cancel appointments first."
        )
    
    await db.execute(
        delete(AgentAvailability).where(AgentAvailability.id == slot_id)
    )
    await db.commit()
    
    return {"status": "deleted", "slot_id": slot_id}


@app.post("/api/tenants/{tenant_id}/schedule")
async def create_schedule_slots(
    tenant_id: int,
    schedule_request: ScheduleSlotsRequest = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """
    Create or replace all schedule slots for a tenant.
    Deletes all existing slot templates and creates new ones.
    Note: This doesn't affect existing Appointment records.
    """
    try:
        logger.info(f"📅 Schedule API: Received request for tenant {tenant_id}")
        logger.info(f"📅 Request payload: {schedule_request.model_dump()}")
        logger.info(f"📅 Number of slots: {len(schedule_request.slots)}")
        logger.info(f"📅 Credentials present: {credentials is not None}")
        
        if not credentials:
            logger.error("❌ No credentials provided - Authorization header missing or invalid")
            raise HTTPException(status_code=401, detail="Missing Authorization header. Use: Authorization: Bearer <token>")
        
        await verify_tenant_access(credentials, tenant_id, db)
    except HTTPException as e:
        logger.error(f"❌ Auth check failed: {e.detail}")
        raise
    except Exception as e:
        logger.error(f"❌ Unexpected error in auth: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Authorization error: {str(e)}")
    
    logger.info(f"✅ Authentication passed for tenant {tenant_id}")
    
    # Validate all slots before deletion
    from datetime import time as dt_time
    for i, slot_data in enumerate(schedule_request.slots):
        try:
            # Validate time format
            start_hour, start_min = map(int, slot_data.start_time.split(':'))
            end_hour, end_min = map(int, slot_data.end_time.split(':'))
            
            if start_hour < 0 or start_hour > 23 or start_min < 0 or start_min > 59:
                raise ValueError("Invalid start time")
            if end_hour < 0 or end_hour > 23 or end_min < 0 or end_min > 59:
                raise ValueError("Invalid end time")
            
            # Validate day of week
            DayOfWeek(slot_data.day_of_week.lower())
        except (ValueError, KeyError) as e:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid slot at index {i}: {str(e)}. Ensure day_of_week is valid (monday-sunday) and times are in HH:MM format."
            )
    
    # Delete all existing slots (bookings are in Appointment table, so safe to delete templates)
    await db.execute(
        delete(AgentAvailability).where(
            AgentAvailability.tenant_id == tenant_id
        )
    )
    
    # Create new slots
    created_slots = []
    
    for slot_data in schedule_request.slots:
        try:
            # Parse time strings
            start_hour, start_min = map(int, slot_data.start_time.split(':'))
            end_hour, end_min = map(int, slot_data.end_time.split(':'))
            
            # Convert string day to enum
            day_enum = DayOfWeek(slot_data.day_of_week.lower())
            
            new_slot = AgentAvailability(
                tenant_id=tenant_id,
                day_of_week=day_enum,
                start_time=dt_time(start_hour, start_min),
                end_time=dt_time(end_hour, end_min),
                is_booked=False
            )
            db.add(new_slot)
            created_slots.append(new_slot)
        except Exception as e:
            logger.error(f"❌ Error creating slot: {str(e)}", exc_info=True)
            logger.error(f"❌ Slot data that failed: {slot_data.model_dump()}")
            await db.rollback()
            raise HTTPException(status_code=400, detail=f"Failed to create slot: {str(e)}")
    
    await db.commit()
    
    logger.info(f"✅ Successfully created {len(created_slots)} schedule slots for tenant {tenant_id}")
    return {
        "status": "success",
        "created_count": len(created_slots),
        "message": f"Created {len(created_slots)} time slots"
    }


# ==================== APPOINTMENT ENDPOINTS ====================

@app.post("/api/tenants/{tenant_id}/appointments", response_model=AppointmentResponse)
async def create_appointment_endpoint(
    tenant_id: int,
    appointment_data: AppointmentCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Create an appointment. Requires authentication."""
    await verify_tenant_access(credentials, tenant_id, db)
    
    # Verify lead belongs to tenant
    result = await db.execute(
        select(Lead).where(
            Lead.tenant_id == tenant_id,
            Lead.id == appointment_data.lead_id
        )
    )
    lead = result.scalar_one_or_none()
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Create appointment
    appointment = Appointment(
        lead_id=appointment_data.lead_id,
        appointment_type=appointment_data.appointment_type,
        scheduled_date=appointment_data.scheduled_date,
        duration_minutes=appointment_data.duration_minutes,
        location=appointment_data.location,
        meeting_link=appointment_data.meeting_link,
        property_reference=appointment_data.property_reference,
        notes=appointment_data.notes
    )
    db.add(appointment)
    
    # Update lead status
    lead.status = LeadStatus.VIEWING_SCHEDULED
    
    await db.commit()
    await db.refresh(appointment)
    
    return appointment


@app.get("/api/tenants/{tenant_id}/appointments", response_model=List[AppointmentResponse])
async def list_appointments(
    tenant_id: int,
    upcoming_only: bool = False,
    db: AsyncSession = Depends(get_db)
):
    """List appointments for a tenant."""
    # Get lead IDs for tenant
    lead_result = await db.execute(
        select(Lead.id).where(Lead.tenant_id == tenant_id)
    )
    lead_ids = [lid for (lid,) in lead_result.fetchall()]
    
    if not lead_ids:
        return []
    
    query = select(Appointment).where(Appointment.lead_id.in_(lead_ids))
    
    if upcoming_only:
        query = query.where(
            Appointment.scheduled_date >= datetime.utcnow(),
            Appointment.is_cancelled == False
        )
    
    query = query.order_by(Appointment.scheduled_date)
    
    result = await db.execute(query)
    return result.scalars().all()


# ==================== EXCEL EXPORT ====================

@app.get("/api/tenants/{tenant_id}/leads/export")
async def export_leads_excel(
    tenant_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Export all leads to Excel (.xlsx) file."""
    # Get all leads for tenant
    result = await db.execute(
        select(Lead).where(Lead.tenant_id == tenant_id).order_by(Lead.created_at.desc())
    )
    leads = result.scalars().all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Define headers
    headers = [
        "ID", "Name", "Phone", "Telegram Username", "Language", "Status",
        "Transaction Type", "Property Type", "Budget Min", "Budget Max",
        "Payment Method", "Purpose", "Residency Need", "Taste Tags",
        "Voice Transcript", "Source", "Created At", "Last Interaction"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="0f1729", end_color="0f1729", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    gold_border = Border(
        left=Side(style='thin', color='D4AF37'),
        right=Side(style='thin', color='D4AF37'),
        top=Side(style='thin', color='D4AF37'),
        bottom=Side(style='thin', color='D4AF37')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = gold_border
    
    # Write data
    for row, lead in enumerate(leads, 2):
        data = [
            lead.id,
            lead.name or "",
            lead.phone or "",
            lead.telegram_username or "",
            lead.language.value if lead.language else "",
            lead.status.value if lead.status else "",
            lead.transaction_type.value if lead.transaction_type else "",
            lead.property_type.value if lead.property_type else "",
            lead.budget_min or "",
            lead.budget_max or "",
            lead.payment_method.value if lead.payment_method else "",
            lead.purpose.value if lead.purpose else "",
            "Yes" if lead.purpose == Purpose.RESIDENCY else "No",
            ", ".join(lead.taste_tags) if lead.taste_tags else "",
            lead.voice_transcript or "",
            lead.source or "",
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
            lead.last_interaction.strftime("%Y-%m-%d %H:%M") if lead.last_interaction else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = gold_border
            cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    column_widths = [8, 20, 15, 20, 10, 15, 15, 15, 12, 12, 15, 15, 15, 30, 40, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    filename = f"leads_export_{tenant_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== DASHBOARD STATS ====================

@app.get("/api/tenants/{tenant_id}/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(
    tenant_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """Get dashboard statistics for a tenant. Requires authentication."""
    # Verify access (authentication check)
    await verify_tenant_access(credentials, tenant_id, db)
    
    # Get all leads
    result = await db.execute(
        select(Lead).where(Lead.tenant_id == tenant_id)
    )
    leads = result.scalars().all()
    
    # Calculate stats
    total_leads = len(leads)
    active_deals = sum(1 for l in leads if l.status in [LeadStatus.NEGOTIATING, LeadStatus.VIEWING_SCHEDULED])
    qualified_leads = sum(1 for l in leads if l.status == LeadStatus.QUALIFIED)
    scheduled_viewings = sum(1 for l in leads if l.status == LeadStatus.VIEWING_SCHEDULED)
    closed_won = sum(1 for l in leads if l.status == LeadStatus.CLOSED_WON)
    
    conversion_rate = (closed_won / total_leads * 100) if total_leads > 0 else 0
    
    # Status breakdown
    leads_by_status = {}
    for status in LeadStatus:
        count = sum(1 for l in leads if l.status == status)
        if count > 0:
            leads_by_status[status.value] = count
    
    # Purpose breakdown
    leads_by_purpose = {}
    for purpose in Purpose:
        count = sum(1 for l in leads if l.purpose == purpose)
        if count > 0:
            leads_by_purpose[purpose.value] = count
    
    return DashboardStats(
        total_leads=total_leads,
        active_deals=active_deals,
        qualified_leads=qualified_leads,
        scheduled_viewings=scheduled_viewings,
        conversion_rate=round(conversion_rate, 1),
        leads_by_status=leads_by_status,
        leads_by_purpose=leads_by_purpose
    )


# ==================== ROI PDF ENDPOINT ====================

@app.get("/api/tenants/{tenant_id}/leads/{lead_id}/roi-pdf")
async def generate_roi_pdf_endpoint(
    tenant_id: int,
    lead_id: int,
    property_value: Optional[float] = None,
    db: AsyncSession = Depends(get_db)
):
    """Generate ROI PDF report for a lead."""
    # Get tenant
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    # Get lead
    result = await db.execute(
        select(Lead).where(Lead.tenant_id == tenant_id, Lead.id == lead_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Generate PDF
    pdf_bytes = await generate_roi_pdf(tenant, lead, property_value)
    
    # Return as download
    filename = f"roi_analysis_{lead.name or lead_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== TELEGRAM WEBHOOK ====================

@app.post("/webhook/telegram/{bot_token}")
async def telegram_webhook(bot_token: str, update: dict, background_tasks: BackgroundTasks):
    """Handle incoming Telegram webhook updates."""
    background_tasks.add_task(handle_telegram_webhook, bot_token, update)
    return {"status": "ok"}


# ==================== WHATSAPP WEBHOOK ====================

@app.get("/webhook/whatsapp")
async def whatsapp_webhook_verify(
    hub_mode: str = Query(None, alias="hub.mode"),
    hub_token: str = Query(None, alias="hub.verify_token"),
    hub_challenge: str = Query(None, alias="hub.challenge"),
    db: AsyncSession = Depends(get_db)
):
    """
    WhatsApp webhook verification endpoint.
    Called by Meta when registering the webhook.
    Verifies against any tenant's verify_token in the database.
    """
    logger.info(f"🔍 Webhook verify request: mode={hub_mode}, token={hub_token[:20]}..., challenge={hub_challenge}")
    
    if hub_mode != "subscribe":
        logger.warning(f"❌ Invalid mode: {hub_mode}")
        raise HTTPException(status_code=403, detail="Invalid mode")
    
    # Check if any tenant has this verify token
    result = await db.execute(
        select(Tenant).where(Tenant.whatsapp_verify_token == hub_token)
    )
    tenant = result.scalar_one_or_none()
    
    if tenant:
        logger.info(f"✅ Token matched tenant {tenant.id}: {tenant.name}")
        return Response(content=hub_challenge, media_type="text/plain")
    
    # Fallback to environment variable for initial setup
    env_token = os.getenv("WHATSAPP_VERIFY_TOKEN")
    logger.info(f"🔑 Checking env token (hash): {hashlib.sha256(env_token.encode()).hexdigest()[:8] if env_token else 'None'}")
    
    if env_token and secrets.compare_digest(hub_token, env_token):
        logger.info(f"✅ Token matched environment variable")
        return Response(content=hub_challenge, media_type="text/plain")
    else:
        # Don't log actual token values - security risk
        logger.error(f"❌ Token mismatch! Verification failed.")
    
    raise HTTPException(status_code=403, detail="Verification failed")


@app.post("/webhook/whatsapp")
async def whatsapp_webhook(payload: dict, background_tasks: BackgroundTasks):
    """
    Handle incoming WhatsApp webhook updates.
    Routes to appropriate tenant handler based on phone_number_id.
    """
    async def process_webhook():
        await whatsapp_bot_manager.handle_webhook(payload)
    
    background_tasks.add_task(process_webhook)
    return {"status": "ok"}


@app.post("/api/webhook/waha")
async def waha_webhook(
    request: Request,
    payload: dict,
    background_tasks: BackgroundTasks,
    x_tenant_id: Optional[int] = Header(None, alias="X-Tenant-ID")
):
    """
    Handle incoming Waha (self-hosted WhatsApp) webhook updates.
    This receives messages from the WhatsApp Router with tenant routing.
    
    Router adds X-Tenant-ID header to indicate which tenant this message belongs to.
    """
    async def process_waha_webhook():
        try:
            logger.info(f"📨 Waha webhook received: {payload.get('event')}")
            
            # Parse webhook using Waha provider
            from whatsapp_providers import WahaWhatsAppProvider
            from database import Tenant
            
            # Get tenant from router header or fallback to default
            async with AsyncSessionLocal() as db:
                if x_tenant_id:
                    logger.info(f"🔀 Routed message for Tenant {x_tenant_id}")
                    result = await db.execute(select(Tenant).where(Tenant.id == x_tenant_id))
                else:
                    logger.warning("⚠️ No X-Tenant-ID header - using default tenant")
                    result = await db.execute(select(Tenant).where(Tenant.id == 1))
                
                tenant = result.scalar_one_or_none()
                
                if not tenant:
                    logger.error(f"❌ Tenant {x_tenant_id} not found")
                    return
                
                provider = WahaWhatsAppProvider(tenant)
                message_data = provider.parse_webhook(payload)
                
                if not message_data:
                    logger.info("Waha webhook ignored (not a user message)")
                    return
                
                # Route to WhatsApp bot manager
                # Convert to Meta-like format for compatibility
                meta_format = {
                    "entry": [{
                        "changes": [{
                            "value": {
                                "messages": [{
                                    "from": message_data["from_phone"],
                    "type": message_data["message_type"],
                                    "text": {"body": message_data.get("text", "")}
                                }],
                                "contacts": [{
                                    "profile": {"name": message_data.get("profile_name", "User")}
                                }]
                            }
                        }]
                    }]
                }
                
                await whatsapp_bot_manager.handle_webhook(meta_format)
                
        except Exception as e:
            logger.error(f"❌ Error processing Waha webhook: {e}", exc_info=True)
    
    background_tasks.add_task(process_waha_webhook)
    return {"status": "ok"}


# ==================== TENANT DATA MANAGEMENT ====================
# These endpoints allow agents to manage their properties, projects, and knowledge base
# The AI uses this data to provide personalized responses to leads

# --- Properties ---

@app.get("/api/tenants/{tenant_id}/properties", response_model=List[TenantPropertyResponse])
async def list_properties(
    tenant_id: int,
    skip: int = 0,
    limit: int = 50,
    property_type: Optional[PropertyType] = None,
    is_available: Optional[bool] = None,
    db: AsyncSession = Depends(get_db)
):
    """List all properties for a tenant. AI uses these for recommendations."""
    query = select(TenantProperty).where(TenantProperty.tenant_id == tenant_id)
    
    if property_type:
        query = query.where(TenantProperty.property_type == property_type)
    if is_available is not None:
        query = query.where(TenantProperty.is_available == is_available)
    
    query = query.order_by(TenantProperty.is_featured.desc(), TenantProperty.created_at.desc())
    query = query.offset(skip).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()


@app.post("/api/tenants/{tenant_id}/properties", response_model=TenantPropertyResponse)
async def create_property(
    tenant_id: int,
    property_data: TenantPropertyCreate,
    db: AsyncSession = Depends(get_db)
):
    """Add a new property to tenant's inventory. AI will use this for recommendations."""
    try:
        # Verify tenant exists
        result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="Tenant not found")
        
        # Extract data and handle enum fields
        property_dict = property_data.model_dump()
        
        # Ensure enums are properly converted to uppercase strings (matching database enum values)
        if 'property_type' in property_dict and property_dict['property_type']:
            if isinstance(property_dict['property_type'], PropertyType):
                property_dict['property_type'] = property_dict['property_type'].value.upper()
            elif isinstance(property_dict['property_type'], str):
                property_dict['property_type'] = property_dict['property_type'].upper()
        
        if 'transaction_type' in property_dict and property_dict['transaction_type']:
            if isinstance(property_dict['transaction_type'], TransactionType):
                property_dict['transaction_type'] = property_dict['transaction_type'].value.upper()
            elif isinstance(property_dict['transaction_type'], str):
                property_dict['transaction_type'] = property_dict['transaction_type'].upper()
        
        property_obj = TenantProperty(
            tenant_id=tenant_id,
            **property_dict
        )
        db.add(property_obj)
        await db.commit()
        await db.refresh(property_obj)
        
        logger.info(f"✅ Property created successfully for tenant {tenant_id}: {property_obj.id}")
        return property_obj
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Failed to create property: {str(e)}", exc_info=True)
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to create property: {str(e)}")


@app.put("/api/tenants/{tenant_id}/properties/{property_id}", response_model=TenantPropertyResponse)
async def update_property(
    tenant_id: int,
    property_id: int,
    property_data: TenantPropertyCreate,
    db: AsyncSession = Depends(get_db)
):
    """
    Update a property in tenant's inventory.
    NOTE: This does NOT update image_urls/image_files - use image upload/delete endpoints for that.
    """
    result = await db.execute(
        select(TenantProperty).where(
            TenantProperty.tenant_id == tenant_id,
            TenantProperty.id == property_id
        )
    )
    property_obj = result.scalar_one_or_none()
    
    if not property_obj:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # اصلاح: حفظ image-related fields و فقط بروزرسانی فیلدهای دیگر
    update_data = property_data.model_dump(exclude={'images'})
    
    for key, value in update_data.items():
        # حفظ image_urls، image_files، primary_image
        if key not in ['image_urls', 'image_files', 'primary_image']:
            setattr(property_obj, key, value)
    
    property_obj.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(property_obj)
    
    return property_obj


@app.delete("/api/tenants/{tenant_id}/properties/{property_id}")
async def delete_property(
    tenant_id: int,
    property_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete a property from tenant's inventory."""
    result = await db.execute(
        select(TenantProperty).where(
            TenantProperty.tenant_id == tenant_id,
            TenantProperty.id == property_id
        )
    )
    property_obj = result.scalar_one_or_none()
    
    if not property_obj:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Delete property images from filesystem
    from file_manager import file_manager
    if property_obj.image_files:
        deleted_count = file_manager.delete_property_images(property_obj.image_files)
        logger.info(f"Deleted {deleted_count} images for property {property_id}")
    
    # Cleanup property directory
    file_manager.cleanup_property_directory(tenant_id, property_id)
    
    await db.delete(property_obj)
    await db.commit()
    
    return {"status": "deleted", "id": property_id}


@app.post("/api/tenants/{tenant_id}/properties/{property_id}/images")
async def upload_property_images(
    tenant_id: int,
    property_id: int,
    files: List[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Upload multiple images for a property (max 5 images, 5MB each).
    Returns list of uploaded file metadata.
    """
    from file_manager import file_manager, MAX_IMAGES_PER_PROPERTY
    
    # Verify property exists
    result = await db.execute(
        select(TenantProperty).where(
            TenantProperty.tenant_id == tenant_id,
            TenantProperty.id == property_id
        )
    )
    property_obj = result.scalar_one_or_none()
    
    if not property_obj:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Check current image count - بررسی تعداد عکس‌ها
    current_images = property_obj.image_files or []
    current_count = len(current_images)
    new_count = len(files)
    total_count = current_count + new_count
    
    # محدودیت: حداکثر 5 عکس
    if total_count > MAX_IMAGES_PER_PROPERTY:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "تعداد عکس‌ها بیش از حد مجاز است",
                "message": f"حداکثر {MAX_IMAGES_PER_PROPERTY} عکس برای هر ملک مجاز است",
                "current": current_count,
                "attempting": new_count,
                "max_allowed": MAX_IMAGES_PER_PROPERTY
            }
        )
    
    # بررسی تعداد فایل‌های ارسالی
    if new_count == 0:
        raise HTTPException(
            status_code=400,
            detail="هیچ فایلی برای آپلود انتخاب نشده است"
        )
    
    if new_count > MAX_IMAGES_PER_PROPERTY:
        raise HTTPException(
            status_code=400,
            detail=f"نمی‌توانید بیش از {MAX_IMAGES_PER_PROPERTY} فایل به یکباره آپلود کنید"
        )
    
    # Upload files - آپلود و بررسی هر فایل
    uploaded_files = []
    failed_files = []
    
    for file in files:
        try:
            # بررسی نوع فایل از MIME type
            if not file.content_type or not file.content_type.startswith('image/'):
                failed_files.append({
                    "filename": file.filename,
                    "error": f"فایل باید عکس باشد. نوع فایل: {file.content_type}"
                })
                continue
            
            # Read file data
            file_data = await file.read()
            
            # بررسی حجم فایل قبل از ذخیره
            file_size_mb = len(file_data) / 1024 / 1024
            if file_size_mb > 3:
                failed_files.append({
                    "filename": file.filename,
                    "error": f"حجم فایل ({file_size_mb:.2f}MB) بیش از حد مجاز (3MB) است"
                })
                continue
            
            # Save file with MIME type validation
            file_metadata = file_manager.save_property_image(
                file_data=file_data,
                filename=file.filename,
                tenant_id=tenant_id,
                property_id=property_id,
                content_type=file.content_type
            )
            uploaded_files.append(file_metadata)
            
        except ValueError as e:
            # Validation error (size, type, etc.)
            failed_files.append({
                "filename": file.filename,
                "error": str(e)
            })
        except Exception as e:
            logger.error(f"Failed to upload {file.filename}: {e}")
            failed_files.append({
                "filename": file.filename,
                "error": f"خطا در آپلود فایل: {str(e)}"
            })
    
    # اگر هیچ فایلی آپلود نشد
    if len(uploaded_files) == 0:
        error_messages = [f"{f['filename']}: {f['error']}" for f in failed_files]
        raise HTTPException(
            status_code=400,
            detail={
                "error": "آپلود فایل‌ها با خطا مواجه شد",
                "failed_files": failed_files,
                "messages": error_messages
            }
        )
    
    # Update property with new images
    all_images = current_images + uploaded_files
    property_obj.image_files = all_images
    property_obj.image_urls = [img["url"] for img in all_images]
    property_obj.images = property_obj.image_urls  # Legacy field
    
    # Set first image as primary if not set
    if not property_obj.primary_image and uploaded_files:
        property_obj.primary_image = uploaded_files[0]["url"]
    
    property_obj.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(property_obj)
    
    # پاسخ با جزئیات کامل
    response = {
        "status": "success",
        "message": f"{len(uploaded_files)} عکس با موفقیت آپلود شد",
        "uploaded": len(uploaded_files),
        "files": uploaded_files,
        "total_images": len(all_images),
        "max_allowed": MAX_IMAGES_PER_PROPERTY,
        "remaining_slots": MAX_IMAGES_PER_PROPERTY - len(all_images)
    }
    
    # اگر برخی فایل‌ها با خطا مواجه شدند
    if failed_files:
        response["warnings"] = {
            "message": f"{len(failed_files)} فایل آپلود نشد",
            "failed_files": failed_files
        }
    
    return response


@app.post("/api/tenants/{tenant_id}/properties/upload-pdf")
async def upload_property_pdf(
    tenant_id: int,
    file: UploadFile = File(...),
    extract_text: bool = Query(False, description="Extract text from PDF for auto-fill"),
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
):
    """
    Upload a property brochure/flyer PDF.
    Optionally extracts text for auto-filling property details.
    
    Returns:
    - file_url: URL to access the PDF
    - extracted_data: Dict with parsed property information (if extract_text=True)
    """
    # Verify tenant access
    await verify_tenant_access(credentials, tenant_id, db)
    
    try:
        logger.info(f"📄 PDF Upload: tenant_id={tenant_id}, filename={file.filename}, content_type={file.content_type}")
        logger.info(f"📄 Extract text flag: {extract_text}")
        
        # Verify tenant exists
        result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
        tenant = result.scalar_one_or_none()
        if not tenant:
            logger.error(f"❌ Tenant {tenant_id} not found for PDF upload")
            raise HTTPException(status_code=404, detail="Tenant not found")
        
        # Validate file type
        if not file.content_type or file.content_type != 'application/pdf':
            logger.error(f"❌ Invalid file type: {file.content_type}")
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type. Expected PDF, got {file.content_type}"
            )
        
        # Read file data
        try:
            file_data = await file.read()
            logger.info(f"✅ PDF file read successfully: {len(file_data)} bytes ({len(file_data)/1024/1024:.2f}MB)")
        except Exception as e:
            logger.error(f"❌ Failed to read PDF file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")
        
        # Validate file size (max 50MB for PDFs - matches nginx limit)
        file_size_mb = len(file_data) / 1024 / 1024
        if file_size_mb > 50:
            logger.error(f"PDF file too large: {file_size_mb:.2f}MB")
            raise HTTPException(
                status_code=400,
                detail=f"File size ({file_size_mb:.2f}MB) exceeds maximum allowed (50MB)"
            )
        
        # Generate safe filename
        import re
        import secrets
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '_', file.filename or 'property')
        unique_filename = f"{tenant_id}_{timestamp}_{secrets.token_hex(4)}_{safe_filename}"
        
        # Save PDF to upload directory
        try:
            pdf_dir = os.path.join(UPLOAD_DIR, "pdfs")
            os.makedirs(pdf_dir, exist_ok=True)
            file_path = os.path.join(pdf_dir, unique_filename)
            
            with open(file_path, "wb") as f:
                f.write(file_data)
            
            logger.info(f"PDF saved successfully: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save PDF file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
        
        # Generate URL
        file_url = f"/uploads/pdfs/{unique_filename}"
        
        response = {
            "status": "success",
            "message": "PDF uploaded successfully",
            "file_url": file_url,
            "filename": unique_filename,
            "size_mb": round(file_size_mb, 2)
        }
        
        # Optional: Extract text from PDF for auto-fill
        if extract_text:
            try:
                import PyPDF2
                logger.info(f"Extracting text from PDF: {file_path}")
                
                # Extract text from PDF
                with open(file_path, "rb") as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    text = ""
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
                
                logger.info(f"Extracted {len(text)} characters from PDF")
                
                # Basic extraction patterns (can be enhanced with AI later)
                extracted_data = {}
                
                # Extract price (AED amounts)
                price_match = re.search(r'AED\s*([\d,]+)', text, re.IGNORECASE)
                if price_match:
                    price_str = price_match.group(1).replace(',', '')
                    extracted_data["price"] = float(price_str)
                
                # Extract bedrooms
                bed_match = re.search(r'(\d+)\s*(bed|bedroom|BR)', text, re.IGNORECASE)
                if bed_match:
                    extracted_data["bedrooms"] = int(bed_match.group(1))
                
                # Extract area (sqft)
                area_match = re.search(r'([\d,]+)\s*(sq\.?\s*ft|sqft|square feet)', text, re.IGNORECASE)
                if area_match:
                    area_str = area_match.group(1).replace(',', '')
                    extracted_data["area_sqft"] = float(area_str)
                
                # Extract location mentions (Dubai areas)
                dubai_areas = [
                    "Downtown Dubai", "Dubai Marina", "Palm Jumeirah", "JBR", "Business Bay",
                    "JLT", "Jumeirah Lake Towers", "Arabian Ranches", "Dubai Hills",
                    "City Walk", "DIFC", "Burj Khalifa", "Emirates Hills", "Meydan"
                ]
                for area in dubai_areas:
                    if area.lower() in text.lower():
                        extracted_data["location"] = area
                        break
                
                response["extracted_data"] = extracted_data
                response["extracted_text_preview"] = text[:500] + "..." if len(text) > 500 else text
                logger.info(f"✅ Text extraction successful: {len(extracted_data)} fields extracted")
                
            except ImportError as e:
                logger.warning(f"⚠️ PyPDF2 not installed: {e}")
                response["warning"] = "PyPDF2 not installed. Text extraction unavailable. Install with: pip install PyPDF2"
            except Exception as e:
                logger.error(f"❌ PDF text extraction failed: {e}", exc_info=True)
                response["warning"] = f"Text extraction failed: {str(e)}"
        
        logger.info(f"✅ PDF upload completed successfully for tenant {tenant_id}")
        return response
        
    except Exception as outer_error:
        logger.error(f"❌ PDF upload failed with unexpected error: {outer_error}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(outer_error)}")


@app.delete("/api/tenants/{tenant_id}/properties/{property_id}/images/{filename}")
async def delete_property_image(
    tenant_id: int,
    property_id: int,
    filename: str,
    db: AsyncSession = Depends(get_db)
):
    """Delete a specific image from a property."""
    from file_manager import file_manager
    
    # Get property
    result = await db.execute(
        select(TenantProperty).where(
            TenantProperty.tenant_id == tenant_id,
            TenantProperty.id == property_id
        )
    )
    property_obj = result.scalar_one_or_none()
    
    if not property_obj:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Find image in metadata
    image_files = property_obj.image_files or []
    image_to_delete = None
    remaining_images = []
    
    for img in image_files:
        if img.get("filename") == filename:
            image_to_delete = img
        else:
            remaining_images.append(img)
    
    if not image_to_delete:
        raise HTTPException(status_code=404, detail="Image not found")
    
    # Delete file from filesystem
    file_manager.delete_property_image(image_to_delete.get("path", ""))
    
    # Update property
    property_obj.image_files = remaining_images
    property_obj.image_urls = [img["url"] for img in remaining_images]
    property_obj.images = property_obj.image_urls
    
    # Update primary image if deleted
    if property_obj.primary_image == image_to_delete.get("url"):
        property_obj.primary_image = remaining_images[0]["url"] if remaining_images else None
    
    property_obj.updated_at = datetime.utcnow()
    await db.commit()
    
    return {"status": "deleted", "filename": filename, "remaining": len(remaining_images)}


# --- Projects ---

@app.get("/api/tenants/{tenant_id}/projects", response_model=List[TenantProjectResponse])
async def list_projects(
    tenant_id: int,
    skip: int = 0,
    limit: int = 50,
    is_active: Optional[bool] = None,
    db: AsyncSession = Depends(get_db)
):
    """List all off-plan projects for a tenant. AI uses these for recommendations."""
    query = select(TenantProject).where(TenantProject.tenant_id == tenant_id)
    
    if is_active is not None:
        query = query.where(TenantProject.is_active == is_active)
    
    query = query.order_by(TenantProject.is_featured.desc(), TenantProject.created_at.desc())
    query = query.offset(skip).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()


@app.post("/api/tenants/{tenant_id}/projects", response_model=TenantProjectResponse)
async def create_project(
    tenant_id: int,
    project_data: TenantProjectCreate,
    db: AsyncSession = Depends(get_db)
):
    """Add a new off-plan project. AI will use this for investment recommendations."""
    # Verify tenant exists
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    project = TenantProject(
        tenant_id=tenant_id,
        **project_data.model_dump()
    )
    db.add(project)
    await db.commit()
    await db.refresh(project)
    
    return project


@app.put("/api/tenants/{tenant_id}/projects/{project_id}", response_model=TenantProjectResponse)
async def update_project(
    tenant_id: int,
    project_id: int,
    project_data: TenantProjectCreate,
    db: AsyncSession = Depends(get_db)
):
    """Update an off-plan project."""
    result = await db.execute(
        select(TenantProject).where(
            TenantProject.tenant_id == tenant_id,
            TenantProject.id == project_id
        )
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    for key, value in project_data.model_dump().items():
        setattr(project, key, value)
    
    project.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(project)
    
    return project


@app.delete("/api/tenants/{tenant_id}/projects/{project_id}")
async def delete_project(
    tenant_id: int,
    project_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete an off-plan project."""
    result = await db.execute(
        select(TenantProject).where(
            TenantProject.tenant_id == tenant_id,
            TenantProject.id == project_id
        )
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    db.delete(project)
    await db.commit()
    
    return {"status": "deleted", "id": project_id}


# --- Knowledge Base ---

@app.get("/api/tenants/{tenant_id}/knowledge", response_model=List[TenantKnowledgeResponse])
async def list_knowledge(
    tenant_id: int,
    category: Optional[str] = None,
    language: Optional[Language] = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db)
):
    """List knowledge base entries. AI uses these for FAQ answers."""
    query = select(TenantKnowledge).where(
        TenantKnowledge.tenant_id == tenant_id,
        TenantKnowledge.is_active == True
    )
    
    if category:
        query = query.where(TenantKnowledge.category == category)
    if language:
        query = query.where(TenantKnowledge.language == language)
    
    query = query.order_by(TenantKnowledge.priority.desc())
    query = query.offset(skip).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()


@app.post("/api/tenants/{tenant_id}/knowledge", response_model=TenantKnowledgeResponse)
async def create_knowledge(
    tenant_id: int,
    knowledge_data: TenantKnowledgeCreate,
    db: AsyncSession = Depends(get_db)
):
    """Add a knowledge base entry. AI will use this for answering questions."""
    # Verify tenant exists
    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Tenant not found")
    
    knowledge = TenantKnowledge(
        tenant_id=tenant_id,
        **knowledge_data.model_dump()
    )
    db.add(knowledge)
    await db.commit()
    await db.refresh(knowledge)
    
    return knowledge


@app.put("/api/tenants/{tenant_id}/knowledge/{knowledge_id}", response_model=TenantKnowledgeResponse)
async def update_knowledge(
    tenant_id: int,
    knowledge_id: int,
    knowledge_data: TenantKnowledgeCreate,
    db: AsyncSession = Depends(get_db)
):
    """Update a knowledge base entry."""
    result = await db.execute(
        select(TenantKnowledge).where(
            TenantKnowledge.tenant_id == tenant_id,
            TenantKnowledge.id == knowledge_id
        )
    )
    knowledge = result.scalar_one_or_none()
    
    if not knowledge:
        raise HTTPException(status_code=404, detail="Knowledge entry not found")
    
    for key, value in knowledge_data.model_dump().items():
        setattr(knowledge, key, value)
    
    knowledge.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(knowledge)
    
    return knowledge


@app.delete("/api/tenants/{tenant_id}/knowledge/{knowledge_id}")
async def delete_knowledge(
    tenant_id: int,
    knowledge_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete a knowledge base entry."""
    result = await db.execute(
        select(TenantKnowledge).where(
            TenantKnowledge.tenant_id == tenant_id,
            TenantKnowledge.id == knowledge_id
        )
    )
    knowledge = result.scalar_one_or_none()
    
    if not knowledge:
        raise HTTPException(status_code=404, detail="Knowledge entry not found")
    
    db.delete(knowledge)
    await db.commit()
    
    return {"status": "deleted", "id": knowledge_id}


# ==================== MAIN ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=int(os.getenv("PORT", "8000")),
        reload=os.getenv("DEBUG", "false").lower() == "true"
    )
